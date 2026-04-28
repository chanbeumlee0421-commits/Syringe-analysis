import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# 페이지 설정
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="주사기 거래재개 병원 분석",
    page_icon="💉",
    layout="wide",
)

st.markdown("""
<style>
    [data-testid="stAppViewContainer"] { background: #f0f4f8; }
    [data-testid="stSidebar"] { background: #1F4E79 !important; }
    [data-testid="stSidebar"] label,
    [data-testid="stSidebar"] p,
    [data-testid="stSidebar"] span,
    [data-testid="stSidebar"] div { color: white !important; }
    .metric-card {
        background: white;
        border-radius: 12px;
        padding: 20px 24px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        border-left: 5px solid #2E75B6;
        margin-bottom: 8px;
    }
    .metric-label { font-size: 13px; color: #666; margin-bottom: 4px; }
    .metric-value { font-size: 28px; font-weight: 700; color: #1F4E79; }
    .metric-sub   { font-size: 12px; color: #999; margin-top: 2px; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# 사이드바
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 💉 분석 설정")
    st.markdown("---")
    uploaded = st.file_uploader("엑셀 파일 업로드", type=["xlsx", "xls"])
    st.markdown("---")
    gap_days = st.slider(
        "공백 기준 (일)",
        min_value=30, max_value=730, value=365, step=30,
        help="비주사기 마지막 주문 후 이 기간 이상 지난 뒤 주사기 주문 시 '재개'로 판단"
    )
    st.caption(f"현재 기준: **{gap_days}일** ({round(gap_days/30, 1)}개월) 이상")
    st.markdown("---")
    st.markdown("**주사기 키워드**")
    keywords_input = st.text_input("쉼표로 구분", value="Syringe,주사기,LDS")
    keywords = [k.strip() for k in keywords_input.split(",") if k.strip()]


# ─────────────────────────────────────────────
# 분석 함수
# ─────────────────────────────────────────────
@st.cache_data
def analyze(file_bytes, gap_threshold, kws):
    df = pd.read_excel(BytesIO(file_bytes))
    st.write(df.columns.tolist())
    df['매출일_date'] = pd.to_datetime(df['매출일(배송완료일)'], errors='coerce')

    pattern      = '|'.join(kws)
    syringe_mask = df['제품명'].str.contains(pattern, na=False)
    direct_mask  = df['유통'] == '직거래'

    syr_df     = df[syringe_mask & direct_mask].copy()
    non_syr_df = df[~syringe_mask & direct_mask].copy()

    if syr_df.empty:
        return pd.DataFrame()

    # 비주사기 마지막 매출일
    last_non = non_syr_df.groupby('거래처명')['매출일_date'].max().reset_index()
    last_non.columns = ['거래처명', '비주사기_마지막매출일']

    # 주사기 집계
    syr_agg = syr_df.groupby('거래처명').agg(
        주사기_첫매출일   = ('매출일_date', 'min'),
        주사기_최근매출일 = ('매출일_date', 'max'),
        주사기_주문횟수   = ('매출일_date', 'count'),
    ).reset_index()

    # 담당자
    담당자s = syr_df.groupby('거래처명')['담당자'].agg(
        lambda x: x.dropna().mode().iloc[0] if len(x.dropna()) > 0 else ''
    ).reset_index()
    syr_agg = syr_agg.merge(담당자s, on='거래처명', how='left')

    # 지역
    지역s = syr_df.groupby('거래처명')[['지역1', '지역2']].first().reset_index()
    syr_agg = syr_agg.merge(지역s, on='거래처명', how='left')

    # 주사기 제품 목록
    prods = syr_df.groupby('거래처명')['제품명'].apply(
        lambda x: ', '.join(sorted(x.unique()))
    ).reset_index()
    prods.columns = ['거래처명', '주문_주사기_제품']
    syr_agg = syr_agg.merge(prods, on='거래처명', how='left')

    # 공백 계산
    result = syr_agg.merge(last_non, on='거래처명', how='left')
    result['공백기간_일']  = (result['주사기_첫매출일'] - result['비주사기_마지막매출일']).dt.days
    result['공백기간_개월'] = (result['공백기간_일'] / 30).round(1)

    comeback = result[result['공백기간_일'] >= gap_threshold].copy()
    comeback = comeback.sort_values('공백기간_일', ascending=False).reset_index(drop=True)
    return comeback


# ─────────────────────────────────────────────
# 엑셀 다운로드 생성 함수
# ─────────────────────────────────────────────
def make_excel(comeback: pd.DataFrame, gap_days: int) -> bytes:
    def fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    def thin_border(color='BDD7EE'):
        s = Side(style='thin', color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    C_HDR = '1F4E79'
    C_FG  = 'FFFFFF'
    C_SUB = 'D6E4F0'
    C_SFG = '1F4E79'
    C_EVN = 'EBF3FB'
    C_ODD = 'FFFFFF'

    wb = Workbook()
    ws = wb.active
    ws.title = '거래재개 병원 목록'
    ws.sheet_view.showGridLines = False

    # 타이틀
    ws.merge_cells('A1:K1')
    ws['A1'].value     = f'주사기 거래 재개 병원 목록  |  공백 {gap_days}일 이상 후 주사기 신규 주문'
    ws['A1'].font      = Font(name='Arial', bold=True, size=13, color=C_FG)
    ws['A1'].fill      = fill(C_HDR)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 34

    # 요약 카드
    n        = len(comeback)
    max_gap  = int(comeback['공백기간_일'].max())
    max_gapm = comeback['공백기간_개월'].max()
    avg_gap  = int(comeback['공백기간_일'].mean())
    avg_gapm = round(comeback['공백기간_개월'].mean(), 1)
    avg_ord  = round(comeback['주사기_주문횟수'].mean(), 1)

    for rng, txt in [
        ('A2:B2', f'총 해당 병원\n{n}개'),
        ('C2:D2', f'최장 공백\n{max_gap}일 ({max_gapm}개월)'),
        ('E2:F2', f'평균 공백\n{avg_gap}일 ({avg_gapm}개월)'),
        ('G2:H2', f'평균 주문 횟수\n{avg_ord}회'),
    ]:
        ws.merge_cells(rng)
        c = ws[rng.split(':')[0]]
        c.value     = txt
        c.font      = Font(name='Arial', bold=True, size=10, color=C_SFG)
        c.fill      = fill(C_SUB)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = thin_border()
    ws.row_dimensions[2].height = 38

    # 헤더
    headers    = ['No.','거래처명','지역','담당자','비주사기\n마지막 주문','주사기\n첫 주문','주사기\n최근 주문','공백(일)','공백(개월)','주사기\n주문횟수','주문한 주사기 제품']
    col_widths = [5, 22, 12, 12, 16, 16, 16, 9, 9, 10, 48]
    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        c = ws.cell(row=3, column=ci, value=h)
        c.font      = Font(name='Arial', bold=True, size=10, color=C_FG)
        c.fill      = fill(C_HDR)
        c.alignment = center
        c.border    = thin_border()
    ws.row_dimensions[3].height = 30

    # 데이터
    def fmt_date(v):
        try:
            return pd.to_datetime(v).strftime('%Y-%m-%d')
        except:
            return ''

    for ri, row in comeback.iterrows():
        er  = ri + 4
        rf  = fill(C_EVN if ri % 2 == 0 else C_ODD)
        gap = row['공백기간_일']
        gf  = fill('FFCCCC') if gap >= 900 else fill('FFE0B2') if gap >= 600 else fill('FFF9C4')

        vals = [
            ri + 1,
            row['거래처명'],
            f"{row['지역1']} {row['지역2']}".strip(),
            row['담당자'],
            fmt_date(row['비주사기_마지막매출일']),
            fmt_date(row['주사기_첫매출일']),
            fmt_date(row['주사기_최근매출일']),
            int(row['공백기간_일']),
            row['공백기간_개월'],
            int(row['주사기_주문횟수']),
            row['주문_주사기_제품'],
        ]
        for ci, val in enumerate(vals, start=1):
            c = ws.cell(row=er, column=ci, value=val)
            c.border    = thin_border()
            c.font      = Font(name='Arial', size=10)
            c.fill      = gf if ci in (8, 9) else rf
            c.font      = Font(name='Arial', bold=ci in (8, 9), size=10)
            c.alignment = left if ci == 11 else center
        ws.row_dimensions[er].height = 20

    # 담당자별 시트
    ws2 = wb.create_sheet('담당자별 요약')
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells('A1:F1')
    ws2['A1'].value     = '담당자별 거래 재개 현황'
    ws2['A1'].font      = Font(name='Arial', bold=True, size=13, color=C_FG)
    ws2['A1'].fill      = fill('2E75B6')
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 32

    hdrs2   = ['담당자','병원 수','평균 공백(일)','평균 공백(개월)','총 주문 횟수','병원 목록']
    widths2 = [14, 10, 14, 14, 12, 55]
    for ci, (h, w) in enumerate(zip(hdrs2, widths2), start=1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
        c = ws2.cell(row=2, column=ci, value=h)
        c.font      = Font(name='Arial', bold=True, size=10, color=C_FG)
        c.fill      = fill(C_HDR)
        c.alignment = center
        c.border    = thin_border()
    ws2.row_dimensions[2].height = 28

    grp = comeback.groupby('담당자').agg(
        병원수=('거래처명','count'),
        평균공백일=('공백기간_일','mean'),
        평균공백개월=('공백기간_개월','mean'),
        총주문횟수=('주사기_주문횟수','sum'),
        병원목록=('거래처명', lambda x: ', '.join(x))
    ).reset_index().sort_values('병원수', ascending=False).reset_index(drop=True)

    for ri, row in grp.iterrows():
        r   = ri + 3
        rf  = fill(C_EVN if ri % 2 == 0 else C_ODD)
        for ci, val in enumerate([
            row['담당자'], int(row['병원수']), int(row['평균공백일']),
            round(row['평균공백개월'], 1), int(row['총주문횟수']), row['병원목록']
        ], start=1):
            c = ws2.cell(row=r, column=ci, value=val)
            c.font      = Font(name='Arial', size=10)
            c.fill      = rf
            c.border    = thin_border()
            c.alignment = left if ci == 6 else center
        ws2.row_dimensions[r].height = 20

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────
# 메인 화면
# ─────────────────────────────────────────────
st.title("💉 주사기 거래재개 병원 분석")
st.caption("비주사기 주문 공백 후 주사기 신규 주문으로 거래를 재개한 병원을 찾아드립니다.")

if uploaded is None:
    st.info("👈 왼쪽 사이드바에서 엑셀 파일을 업로드해 주세요.")
    st.stop()

with st.spinner("분석 중..."):
    comeback = analyze(uploaded.read(), gap_days, keywords)

if comeback.empty:
    st.warning(f"조건에 맞는 병원이 없어요. 공백 기준({gap_days}일)을 낮춰보세요.")
    st.stop()

# ── 요약 카드 ──
st.markdown("### 📊 요약")
c1, c2, c3, c4 = st.columns(4)

def metric_card(col, label, value, sub=""):
    col.markdown(f"""
    <div class="metric-card">
        <div class="metric-label">{label}</div>
        <div class="metric-value">{value}</div>
        <div class="metric-sub">{sub}</div>
    </div>
    """, unsafe_allow_html=True)

metric_card(c1, "총 해당 병원 수",      f"{len(comeback)}개",                        f"공백 {gap_days}일 이상 기준")
metric_card(c2, "최장 공백",            f"{int(comeback['공백기간_일'].max())}일",    f"{comeback['공백기간_개월'].max()}개월")
metric_card(c3, "평균 공백",            f"{int(comeback['공백기간_일'].mean())}일",   f"{round(comeback['공백기간_개월'].mean(),1)}개월")
metric_card(c4, "평균 주사기 주문 횟수", f"{round(comeback['주사기_주문횟수'].mean(),1)}회", "")

st.markdown("<br>", unsafe_allow_html=True)

# ── 탭 ──
tab1, tab2, tab3 = st.tabs(["📋 전체 목록", "👤 담당자별", "🔍 병원 상세"])

# ── 탭1: 전체 목록 ──
with tab1:
    f1, f2 = st.columns(2)
    with f1:
        담당자_list = ['전체'] + sorted(comeback['담당자'].dropna().unique().tolist())
        sel_담당자 = st.selectbox("담당자 필터", 담당자_list)
    with f2:
        지역_list = ['전체'] + sorted(comeback['지역1'].dropna().unique().tolist())
        sel_지역 = st.selectbox("지역 필터", 지역_list)

    filtered = comeback.copy()
    if sel_담당자 != '전체':
        filtered = filtered[filtered['담당자'] == sel_담당자]
    if sel_지역 != '전체':
        filtered = filtered[filtered['지역1'] == sel_지역]

    display = filtered[[
        '거래처명', '지역1', '지역2', '담당자',
        '비주사기_마지막매출일', '주사기_첫매출일', '주사기_최근매출일',
        '공백기간_일', '공백기간_개월', '주사기_주문횟수', '주문_주사기_제품'
    ]].copy()

    for col in ['비주사기_마지막매출일', '주사기_첫매출일', '주사기_최근매출일']:
        display[col] = pd.to_datetime(display[col]).dt.strftime('%Y-%m-%d')

    display.index = range(1, len(display) + 1)
    display.columns = [
        '거래처명', '지역', '시군구', '담당자',
        '비주사기 마지막 주문', '주사기 첫 주문', '주사기 최근 주문',
        '공백(일)', '공백(개월)', '주문횟수', '주문한 주사기 제품'
    ]

    st.dataframe(
        display,
        use_container_width=True,
        height=500,
        column_config={
            '공백(일)':  st.column_config.NumberColumn(format="%d일"),
            '공백(개월)': st.column_config.NumberColumn(format="%.1f개월"),
            '주문횟수':  st.column_config.NumberColumn(format="%d회"),
        }
    )
    st.caption(f"총 {len(filtered)}개 병원")

# ── 탭2: 담당자별 ──
with tab2:
    grp = comeback.groupby('담당자').agg(
        병원수=('거래처명', 'count'),
        평균공백일=('공백기간_일', 'mean'),
        평균공백개월=('공백기간_개월', 'mean'),
        총주문횟수=('주사기_주문횟수', 'sum'),
        병원목록=('거래처명', lambda x: ', '.join(x))
    ).reset_index().sort_values('병원수', ascending=False)

    grp['평균공백일']   = grp['평균공백일'].round(0).astype(int)
    grp['평균공백개월'] = grp['평균공백개월'].round(1)
    grp.index = range(1, len(grp) + 1)
    grp.columns = ['담당자', '병원 수', '평균 공백(일)', '평균 공백(개월)', '총 주문 횟수', '병원 목록']

    st.dataframe(grp, use_container_width=True, height=400)

    # 바 차트
    st.markdown("**담당자별 병원 수**")
    chart_data = grp[['담당자', '병원 수']].set_index('담당자')
    st.bar_chart(chart_data)

# ── 탭3: 병원 상세 ──
with tab3:
    sel_hospital = st.selectbox("병원 선택", comeback['거래처명'].tolist())
    row = comeback[comeback['거래처명'] == sel_hospital].iloc[0]

    col_a, col_b = st.columns(2)
    with col_a:
        st.markdown("**기본 정보**")
        st.markdown(f"- **지역:** {row['지역1']} {row['지역2']}")
        st.markdown(f"- **담당자:** {row['담당자']}")
        st.markdown(f"- **비주사기 마지막 주문:** {pd.to_datetime(row['비주사기_마지막매출일']).strftime('%Y-%m-%d')}")
    with col_b:
        st.markdown("**주사기 거래 현황**")
        st.markdown(f"- **주사기 첫 주문:** {pd.to_datetime(row['주사기_첫매출일']).strftime('%Y-%m-%d')}")
        st.markdown(f"- **주사기 최근 주문:** {pd.to_datetime(row['주사기_최근매출일']).strftime('%Y-%m-%d')}")
        st.markdown(f"- **공백 기간:** {int(row['공백기간_일'])}일 ({row['공백기간_개월']}개월)")
        st.markdown(f"- **총 주문 횟수:** {int(row['주사기_주문횟수'])}회")

    st.markdown("**주문한 주사기 제품**")
    for p in row['주문_주사기_제품'].split(', '):
        st.markdown(f"  - {p}")

# ── 엑셀 다운로드 ──
st.markdown("---")
excel_bytes = make_excel(comeback, gap_days)
st.download_button(
    label="📥 엑셀 다운로드",
    data=excel_bytes,
    file_name="주사기_거래재개_병원정리.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
